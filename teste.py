from openpyxl import load_workbook

try:
    arquivo_origem = r"C:\Users\USER\Downloads\ArquivoDBF\planilha1.xlsx"
    arquivo_destino = r"C:\Users\USER\Downloads\ArquivoDBF\planilha2.xlsx"

    nome_coluna = "DT_NOTIFIC"

    # Carregar o arquivo de origem
    wb_origem = load_workbook(arquivo_origem)
    ws_origem = wb_origem.active

    # Localizar a coluna pelo cabeçalho
    col_index = None
    for col in range(1, ws_origem.max_column + 1):
        if ws_origem.cell(row=1, column=col).value == nome_coluna:
            col_index = col
            break

    if not col_index:
        raise ValueError(f"Coluna '{nome_coluna}' não encontrada na planilha de origem.")

    # Extrair valores da coluna (mantendo datas no formato original)
    valores = [ws_origem.cell(row=row, column=col_index).value for row in range(2, ws_origem.max_row + 1)]

    # Carregar ou criar o arquivo de destino
    try:
        wb_destino = load_workbook(arquivo_destino)
        ws_destino = wb_destino.active
    except FileNotFoundError:
        wb_destino = load_workbook(arquivo_origem)  # Cria baseado na origem
        ws_destino = wb_destino.active
        ws_destino.delete_rows(2, ws_destino.max_row)  # Remove dados antigos

    # Inserir os valores no destino (na primeira coluna)
    for idx, valor in enumerate(valores, start=2):
        ws_destino.cell(row=idx, column=1, value=valor)

    # Salvar o arquivo de destino
    wb_destino.save(arquivo_destino)
    print(f"Dados da coluna '{nome_coluna}' copiados preservando formato original para '{arquivo_destino}'.")

except Exception as e:
    print(f"Erro: {e}")
