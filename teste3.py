import os
import time
import pandas as pd
from openpyxl import load_workbook

# Arquivos e planilhas
arquivo_origem = r"C:\Users\USER\Downloads\ArquivoDBF\planilha1.xlsx"
planilha_origem = "Planilha1"
arquivo_destino = r"C:\Users\USER\Downloads\ArquivoDBF\planilha2.xlsx"
planilha_destino = "Planilha3"

os.startfile(arquivo_origem)
time.sleep(3)
os.startfile(arquivo_destino)
time.sleep(3)

# Colunas desejadas
colunas_desejadas = ['DT_NOTIFIC', 'SEM_NOT', 'NM_BAIRRO', 'CLASSI_FIN']

# Lê a planilha de origem
df_origem = pd.read_excel(arquivo_origem, sheet_name=planilha_origem)

# Verifica se as colunas desejadas existem na planilha de origem
for coluna in colunas_desejadas:
    if coluna not in df_origem.columns:
        raise ValueError(f"A coluna '{coluna}' não foi encontrada na planilha de origem.")

# Cria um DataFrame com as colunas desejadas
df_dados_copiados = df_origem[colunas_desejadas]

# Formata a coluna DT_NOTIFIC no formato dd/MM/yyyy
df_dados_copiados['DT_NOTIFIC'] = pd.to_datetime(df_dados_copiados['DT_NOTIFIC'], errors='coerce').dt.strftime('%d/%m/%Y')

# Abre a planilha de destino com openpyxl
wb = load_workbook(arquivo_destino)
ws = wb[planilha_destino]

# Localiza as colunas correspondentes na planilha de destino
header = [cell.value for cell in ws[1]]  # Lê o cabeçalho da primeira linha
for coluna in colunas_desejadas:
    if coluna not in header:
        raise ValueError(f"A coluna '{coluna}' não foi encontrada na planilha de destino.")

# Colar os dados nas colunas correspondentes
for index, row in df_dados_copiados.iterrows():
    for coluna in colunas_desejadas:
        col_idx = header.index(coluna) + 1  # Índice da coluna no Excel (base 1)
        ws.cell(row=index + 2, column=col_idx, value=row[coluna])  # Pula o cabeçalho

# Salva as alterações
wb.save(arquivo_destino)
print("Dados colados com sucesso nas colunas correspondentes!")
